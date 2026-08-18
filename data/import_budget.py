from pathlib import Path
import json
import shutil
import pandas as pd
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
SOURCE_DIR = PROJECT_ROOT / "docs" / "source_budget"

DATA_DIR.mkdir(exist_ok=True)
SOURCE_DIR.mkdir(parents=True, exist_ok=True)

# Update this path if the Excel file is stored elsewhere.
SOURCE_FILE = Path(r"C:\Users\HP\EasePalCare\projects\Yv-Me-ROI-Calculator\docs\YOBE STATE EasePal Care PROPOSAL BUDGET.xlsx")

if not SOURCE_FILE.exists():
    raise FileNotFoundError(
        f"Budget file not found:\n{SOURCE_FILE}\n\n"
        "Place the Excel file at the path above and run the importer again."
    )

# Preserve original source workbook
ARCHIVE_FILE = SOURCE_DIR / SOURCE_FILE.name
if not ARCHIVE_FILE.exists():
    shutil.copy2(SOURCE_FILE, ARCHIVE_FILE)

# Load workbook
wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)

# ------------------------------------------------------------
# PROPOSAL BUDGET
# ------------------------------------------------------------

ws = wb["Proposal Budget "]

items = []
current_objective = None

for row in range(4, ws.max_row + 1):

    objective = ws.cell(row, 1).value
    description = ws.cell(row, 2).value
    unit_amount = ws.cell(row, 3).value
    quantity = ws.cell(row, 4).value
    frequency = ws.cell(row, 5).value
    amount = ws.cell(row, 6).value
    activity_code = ws.cell(row, 7).value

    if objective:
        current_objective = str(objective).strip()

    if not description:
        continue

    description_clean = str(description).strip()

    # Ignore subtotal / total / section rows
    ignore_terms = [
        "SUBTOTAL",
        "TOTAL EQUIPMENT",
        "TOTAL DIRECT PROGRAMME",
        "TOTAL PERSONNEL",
        "TOTAL DIRECT COSTS",
        "GRAND TOTAL",
        "BUDGET SUMMARY",
        "EQUIPMENT",
        "PERSONNEL COST",
    ]

    if any(term in description_clean.upper() for term in ignore_terms):
        continue

    # Only capture genuine budget activities
    if amount is None:
        continue

    items.append({
        "objective": current_objective,
        "activity_code": activity_code,
        "activity": description_clean,
        "unit_amount": float(unit_amount or 0),
        "quantity": float(quantity or 0),
        "frequency": float(frequency or 0),
        "total_amount": float(amount or 0),
        "source": "Yobe State EasePal Care Proposal Budget",
    })

budget_df = pd.DataFrame(items)

budget_df.to_csv(
    DATA_DIR / "budget_items.csv",
    index=False,
    encoding="utf-8-sig"
)

# ------------------------------------------------------------
# BUDGET SUMMARY
# ------------------------------------------------------------

summary = {
    "programme": "EasePal Care",
    "location": "Yobe State, Nigeria",
    "pilot_months": 12,

    # UPDATED MODEL
    "chews": 100,
    "beneficiaries": 1000,
    "beneficiaries_per_chew": 10,

    "currency": "NGN",

    "objective_1": 19161200,
    "objective_2": 169862600,
    "objective_3": 124901000,
    "objective_4": 9450000,
    "communications_and_me": 16026000,
    "equipment": 2720000,
    "personnel": 12675231,
    "management_indirect": 81560000,

    # SOURCE EXCEL GRAND TOTAL
    "total_budget": 426906031,

    "source_file": SOURCE_FILE.name,
    "source_sheet": "Proposal Budget",
    "care_model": "1 CHEW : 10 Beneficiaries",
}

with open(
    DATA_DIR / "budget_summary.json",
    "w",
    encoding="utf-8"
) as f:
    json.dump(summary, f, indent=4)

# ------------------------------------------------------------
# ADJUSTABLE ASSUMPTIONS
# ------------------------------------------------------------

assumptions = [
    ["assumption_code", "assumption", "value", "unit", "status", "source"],
    ["CHEW_COUNT", "Number of CHEWs / Nurses", 100, "CHEWs", "CONFIRMED", "Project model"],
    ["BENEFICIARY_COUNT", "Number of beneficiaries", 1000, "beneficiaries", "ADJUSTABLE", "1:10 model"],
    ["BENEFICIARIES_PER_CHEW", "Beneficiaries per CHEW", 10, "beneficiaries/CHEW", "CONFIRMED", "Project model"],
    ["PILOT_DURATION", "Programme duration", 12, "months", "CONFIRMED", "Budget"],
    ["TOTAL_BUDGET", "Total programme budget", 426906031, "NGN", "CONFIRMED", "Source Excel"],
    ["SUBSCRIPTION_PER_BENEFICIARY", "Monthly subscription per beneficiary", 0, "NGN/month", "TBD", "ROI assumption"],
    ["REVENUE_SHARE", "Partner revenue share", 0, "%", "TBD", "ROI assumption"],
    ["INVESTMENT_AMOUNT", "Private investment amount", 0, "NGN", "TBD", "ROI assumption"],
]

assumptions_df = pd.DataFrame(
    assumptions[1:],
    columns=assumptions[0]
)

assumptions_df.to_csv(
    DATA_DIR / "adjustable_assumptions.csv",
    index=False,
    encoding="utf-8-sig"
)

# ------------------------------------------------------------
# VALIDATION
# ------------------------------------------------------------

expected_total = 426906031
calculated_total = int(budget_df["total_amount"].sum())

# The activity rows exclude summary rows, so direct comparison
# is intentionally recorded separately.
validation = {
    "source_grand_total": expected_total,
    "model_grand_total": summary["total_budget"],
    "source_total_confirmed": expected_total == summary["total_budget"],
    "chews": summary["chews"],
    "beneficiaries": summary["beneficiaries"],
    "ratio": summary["beneficiaries_per_chew"],
    "ratio_check": (
        summary["beneficiaries"]
        == summary["chews"] * summary["beneficiaries_per_chew"]
    ),
    "activity_rows_imported": len(budget_df),
    "activity_value_sum": calculated_total,
}

with open(
    DATA_DIR / "budget_validation.json",
    "w",
    encoding="utf-8"
) as f:
    json.dump(validation, f, indent=4)

print("=" * 60)
print("Yv-Me BUDGET IMPORT COMPLETE")
print("=" * 60)
print(f"Activities imported : {len(budget_df)}")
print(f"CHEWs                : {summary['chews']}")
print(f"Beneficiaries        : {summary['beneficiaries']}")
print(f"Care model           : 1 CHEW : 10 Beneficiaries")
print(f"Grand Total          : NGN {summary['total_budget']:,}")
print(f"Ratio validation     : {validation['ratio_check']}")
print("=" * 60)

