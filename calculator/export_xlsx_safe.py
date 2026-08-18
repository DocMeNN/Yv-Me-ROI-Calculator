from pathlib import Path
from openpyxl import Workbook

root = Path(__file__).resolve().parents[1]
output = root / "exports" / "xlsx" / "Yv-Me_ROI_Calculator.xlsx"

wb = Workbook()

ws = wb.active
ws.title = "Executive Summary"

rows = [
    ["Yv-Me ROI Calculator"],
    ["EasePal Care — Yobe State, Nigeria"],
    [],
    ["Care Delivery Model", "1 CHEW : 10 Beneficiaries"],
    ["CHEWs", 100],
    ["Beneficiaries", 1000],
    ["Programme Duration (Months)", 12],
    ["Programme Budget (NGN)", 426906031],
    [],
    ["Purpose", "Programme economics, ROI and scaling analysis"],
]

for row in rows:
    ws.append(row)

ws = wb.create_sheet("Editable Assumptions")

rows = [
    ["ASSUMPTION", "VALUE"],
    ["CHEWs", 100],
    ["Beneficiaries per CHEW", 10],
    ["Programme Duration", 12],
    ["Programme Budget", 426906031],
    ["Subscription / Beneficiary / Month", 15000],
    ["Investment", 426906031],
    ["Revenue Share (%)", 10],
]

for row in rows:
    ws.append(row)

ws = wb.create_sheet("ROI Analysis")

rows = [
    ["METRIC", "VALUE"],
    ["Programme Budget", 426906031],
    ["Monthly Cost", 35575502.58],
    ["Annual Cost", 426906031],
    ["Cost / CHEW", 4269060.31],
    ["Cost / Beneficiary", 426906.031],
    ["Annual Revenue", 180000000],
    ["Annual Contribution", -246906031],
]

for row in rows:
    ws.append(row)

ws = wb.create_sheet("Scaling Analysis")

ws.append([
    "CHEWs",
    "Beneficiaries",
    "Annual Revenue"
])

for chews in [10, 50, 100, 200, 500, 1000, 2000, 5000]:
    ws.append([
        chews,
        chews * 10,
        chews * 10 * 15000 * 12
    ])

ws = wb.create_sheet("Presentation Notes")

notes = [
    ["Yv-Me ROI Calculator"],
    [],
    ["Care Model", "1 CHEW : 10 Beneficiaries"],
    ["Location", "Yobe State, Nigeria"],
    ["Programme Budget", "NGN 426,906,031"],
    [],
    ["Investor Use", "Partnership and investment scenario analysis"],
    ["Donor Use", "Programme funding and cost analysis"],
    ["Grant Use", "Transparent programme economics"],
    ["iNGO Use", "Scale and sustainability analysis"],
]

for row in notes:
    ws.append(row)

for sheet in wb.worksheets:
    for column in sheet.columns:
        max_length = 0

        for cell in column:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        sheet.column_dimensions[
            column[0].column_letter
        ].width = min(max_length + 4, 50)

wb.save(output)

print("XLSX CREATED SUCCESSFULLY")
print(output)
