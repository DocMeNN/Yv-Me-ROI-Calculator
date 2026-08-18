from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from calculator.analytics import calculate

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "exports" / "xlsx" / "Yv-Me_ROI_Calculator.xlsx"

CHEWS = 100
RATIO = 10
MONTHS = 12
BUDGET = 426906031
SUBSCRIPTION = 15000
INVESTMENT = 426906031
REVENUE_SHARE = 10

r = calculate(
    chews=CHEWS,
    beneficiaries_per_chew=RATIO,
    programme_months=MONTHS,
    total_budget=BUDGET,
    subscription_per_beneficiary=SUBSCRIPTION,
    investment_amount=INVESTMENT,
    revenue_share=REVENUE_SHARE,
)

wb = Workbook()

# EXECUTIVE SUMMARY
ws = wb.active
ws.title = "Executive Summary"

data = [
    ["Yv-Me ROI Calculator"],
    ["EasePal Care — Yobe State, Nigeria"],
    [],
    ["Care Delivery Model", "1 CHEW : 10 Beneficiaries"],
    ["CHEWs", CHEWS],
    ["Beneficiaries", r["beneficiaries"]],
    ["Programme Duration", MONTHS],
    ["Programme Budget", BUDGET],
    [],
    ["Monthly Cost", r["monthly_operating_cost"]],
    ["Annual Cost", r["annual_operating_cost"]],
    ["Cost / CHEW", r["cost_per_chew"]],
    ["Cost / Beneficiary", r["cost_per_beneficiary"]],
    ["Monthly Revenue", r["monthly_gross_revenue"]],
    ["Annual Revenue", r["annual_gross_revenue"]],
    ["Annual Contribution", r["annual_contribution"]],
    ["ROI", r["roi"]],
    ["Break-even Months", r["break_even_months"]],
]

for row in data:
    ws.append(row)

# EDITABLE ASSUMPTIONS
ws = wb.create_sheet("Editable Assumptions")

rows = [
    ["ASSUMPTION", "VALUE"],
    ["CHEWs", CHEWS],
    ["Beneficiaries per CHEW", RATIO],
    ["Programme Duration (Months)", MONTHS],
    ["Programme Budget (NGN)", BUDGET],
    ["Subscription / Beneficiary / Month", SUBSCRIPTION],
    ["Investment (NGN)", INVESTMENT],
    ["Revenue Share (%)", REVENUE_SHARE],
]

for row in rows:
    ws.append(row)

# ROI ANALYSIS
ws = wb.create_sheet("ROI Analysis")

rows = [
    ["METRIC", "VALUE"],
    ["CHEWs", CHEWS],
    ["Beneficiaries", r["beneficiaries"]],
    ["Monthly Cost", r["monthly_operating_cost"]],
    ["Annual Cost", r["annual_operating_cost"]],
    ["Monthly Revenue", r["monthly_gross_revenue"]],
    ["Annual Revenue", r["annual_gross_revenue"]],
    ["Annual Contribution", r["annual_contribution"]],
    ["ROI", r["roi"]],
    ["Break-even Months", r["break_even_months"]],
]

for row in rows:
    ws.append(row)

# SCALING
ws = wb.create_sheet("Scaling Analysis")

ws.append([
    "CHEWs",
    "Beneficiaries",
    "Monthly Revenue",
    "Annual Revenue",
    "Annual Contribution",
    "ROI"
])

for scale in [10, 50, 100, 200, 500, 1000, 2000, 5000]:

    s = calculate(
        chews=scale,
        beneficiaries_per_chew=RATIO,
        programme_months=MONTHS,
        total_budget=BUDGET,
        subscription_per_beneficiary=SUBSCRIPTION,
        investment_amount=INVESTMENT,
        revenue_share=REVENUE_SHARE,
    )

    ws.append([
        scale,
        s["beneficiaries"],
        s["monthly_gross_revenue"],
        s["annual_gross_revenue"],
        s["annual_contribution"],
        s["roi"],
    ])

# EXPLANATION
ws = wb.create_sheet("Presentation Notes")

notes = [
    ["Yv-Me ROI Calculator — Presentation Notes"],
    [],
    ["Purpose",
     "Financial model for programme costing, revenue, ROI and scaling analysis."],
    ["Care Model",
     "1 CHEW supports 10 beneficiaries."],
    ["Budget",
     "Based on the imported Yobe State EasePal Care proposal budget."],
    ["Assumptions",
     "Key assumptions can be changed for scenario analysis."],
    ["Investor Use",
     "Supports private partnership and investment discussions."],
    ["Donor Use",
     "Supports programme cost and funding discussions."],
    ["iNGO Use",
     "Supports programme scale and sustainability analysis."],
]

for row in notes:
    ws.append(row)

# FORMATTING
for ws in wb.worksheets:

    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(vertical="center")

    for column in range(1, ws.max_column + 1):

        width = 20

        for row in range(1, ws.max_row + 1):

            value = ws.cell(row, column).value

            if value is not None:
                width = max(
                    width,
                    min(len(str(value)) + 4, 50)
                )

        ws.column_dimensions[
            get_column_letter(column)
        ].width = width

wb.save(OUTPUT)

print("XLSX EXPORT SUCCESSFUL")
print(OUTPUT)
